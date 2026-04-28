"""
Módulo 3 — Generación del Excel
Monitor Económico MX

Recibe los dos DataFrames de procesamiento.py y genera un .xlsx con:
  - Hoja 1 "Resumen": indicadores del día con variaciones y colores
  - Hoja 2 "Histórico": serie diaria USD/MXN últimos 30 días

Uso:
  Importado: from excel_builder import generar_excel
             ruta = generar_excel(df_resumen, df_historico)
"""

import os
from datetime import date
from pathlib import Path

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import (
    Alignment, Border, Font, PatternFill, Side
)
from openpyxl.utils import get_column_letter


# ── Paleta de colores ──────────────────────────────────────────────────────

COLOR = {
    "header_bg":    "1F3864",   # Azul marino — fondo encabezados
    "header_fg":    "FFFFFF",   # Blanco — texto encabezados
    "title_bg":     "2E75B6",   # Azul medio — título principal
    "title_fg":     "FFFFFF",
    "positivo_bg":  "E2EFDA",   # Verde claro — variación positiva
    "positivo_fg":  "375623",   # Verde oscuro
    "negativo_bg":  "FCE4D6",   # Rojo claro — variación negativa
    "negativo_fg":  "843C0C",   # Rojo oscuro
    "neutro_bg":    "F2F2F2",   # Gris claro — sin variación / N/A
    "neutro_fg":    "595959",
    "fila_par":     "DEEAF1",   # Azul muy claro — filas pares
    "fila_impar":   "FFFFFF",   # Blanco — filas impares
    "borde":        "BDD7EE",
}


# ── Helpers de estilo ──────────────────────────────────────────────────────

def _fill(hex_color: str) -> PatternFill:
    return PatternFill("solid", start_color=hex_color, fgColor=hex_color)


def _font(hex_color: str = "000000", bold: bool = False, size: int = 11) -> Font:
    return Font(name="Arial", color=hex_color, bold=bold, size=size)


def _border() -> Border:
    lado = Side(style="thin", color=COLOR["borde"])
    return Border(left=lado, right=lado, top=lado, bottom=lado)


def _centro() -> Alignment:
    return Alignment(horizontal="center", vertical="center", wrap_text=True)


def _izquierda() -> Alignment:
    return Alignment(horizontal="left", vertical="center", wrap_text=True)


def _color_variacion(valor: float | None) -> tuple[str, str]:
    """Devuelve (bg, fg) según si la variación es positiva, negativa o nula."""
    if valor is None:
        return COLOR["neutro_bg"], COLOR["neutro_fg"]
    if valor > 0:
        return COLOR["positivo_bg"], COLOR["positivo_fg"]
    if valor < 0:
        return COLOR["negativo_bg"], COLOR["negativo_fg"]
    return COLOR["neutro_bg"], COLOR["neutro_fg"]


def _formatear_variacion(valor: float | None) -> str:
    """Convierte un float de variación a string legible: +1.23% o N/A."""
    if valor is None:
        return "N/A"
    signo = "+" if valor >= 0 else ""
    return f"{signo}{valor:.2f}%"


def _aplicar_celda(
    ws, fila: int, col: int, valor,
    bg: str = "FFFFFF", fg: str = "000000",
    bold: bool = False, size: int = 11,
    alineacion=None, formato: str = None
):
    """Escribe valor en celda y aplica estilo completo."""
    celda = ws.cell(row=fila, column=col, value=valor)
    celda.fill      = _fill(bg)
    celda.font      = _font(fg, bold=bold, size=size)
    celda.border    = _border()
    celda.alignment = alineacion or _centro()
    if formato:
        celda.number_format = formato
    return celda


# ── Hoja 1: Resumen ────────────────────────────────────────────────────────

def _construir_hoja_resumen(ws, df: pd.DataFrame, fecha_str: str):
    """Construye la hoja de resumen del día."""

    # ── Título principal
    ws.merge_cells("A1:H1")
    celda_titulo = ws["A1"]
    celda_titulo.value     = f"Monitor Económico MX — {fecha_str}"
    celda_titulo.fill      = _fill(COLOR["title_bg"])
    celda_titulo.font      = _font(COLOR["title_fg"], bold=True, size=14)
    celda_titulo.alignment = _centro()
    ws.row_dimensions[1].height = 30

    # ── Subtítulo fuente
    ws.merge_cells("A2:H2")
    celda_sub = ws["A2"]
    celda_sub.value     = "Fuentes: Banco de México (Banxico) · INEGI"
    celda_sub.fill      = _fill(COLOR["fila_par"])
    celda_sub.font      = _font(COLOR["neutro_fg"], size=10)
    celda_sub.alignment = _centro()

    # ── Encabezados de columna
    encabezados = [
        "Indicador", "Descripción", "Valor actual", "Unidad",
        "Fecha dato", "Var. diaria", "Var. semanal", "Var. mensual"
    ]
    anchos = [16, 36, 14, 20, 14, 13, 13, 13]

    for col_idx, (enc, ancho) in enumerate(zip(encabezados, anchos), start=1):
        _aplicar_celda(
            ws, 3, col_idx, enc,
            bg=COLOR["header_bg"], fg=COLOR["header_fg"],
            bold=True, size=11, alineacion=_centro()
        )
        ws.column_dimensions[get_column_letter(col_idx)].width = ancho

    ws.row_dimensions[3].height = 22

    # ── Filas de datos
    for i, row in enumerate(df.itertuples(), start=4):
        bg_fila = COLOR["fila_par"] if i % 2 == 0 else COLOR["fila_impar"]

        # Columnas estáticas
        _aplicar_celda(ws, i, 1, row.indicador,   bg=bg_fila, alineacion=_izquierda())
        _aplicar_celda(ws, i, 2, row.descripcion, bg=bg_fila, alineacion=_izquierda())

        # Valor actual — formato según unidad
        fmt_valor = '#,##0.4f' if row.indicador == "usd_fix" else '#,##0.00'
        _aplicar_celda(ws, i, 3, row.valor, bg=bg_fila, formato=fmt_valor)

        _aplicar_celda(ws, i, 4, row.unidad,    bg=bg_fila, alineacion=_izquierda())
        _aplicar_celda(ws, i, 5, row.fecha_dato, bg=bg_fila)

        # Variaciones con color semafórico
        for col_idx, var_val in enumerate(
            [row.var_diaria, row.var_semanal, row.var_mensual], start=6
        ):
            # Para USD: negativo = peso se fortalece (bueno), invertimos color
            if row.indicador == "usd_fix" and var_val is not None:
                bg_var, fg_var = _color_variacion(-var_val)
            else:
                bg_var, fg_var = _color_variacion(var_val)

            _aplicar_celda(
                ws, i, col_idx,
                _formatear_variacion(var_val),
                bg=bg_var, fg=fg_var, bold=(var_val is not None)
            )

        ws.row_dimensions[i].height = 20

    # ── Nota al pie
    fila_nota = len(df) + 5
    ws.merge_cells(f"A{fila_nota}:H{fila_nota}")
    celda_nota = ws[f"A{fila_nota}"]
    celda_nota.value = (
        "Nota: Verde = favorable para el consumidor MX · "
        "Rojo = desfavorable · N/A = dato no disponible para ese período"
    )
    celda_nota.font      = _font(COLOR["neutro_fg"], size=9)
    celda_nota.alignment = _izquierda()

    # Freeze encabezados
    ws.freeze_panes = "A4"


# ── Hoja 2: Histórico ──────────────────────────────────────────────────────

def _construir_hoja_historico(ws, df: pd.DataFrame):
    """Construye la hoja de histórico del tipo de cambio."""

    # ── Título
    ws.merge_cells("A1:E1")
    celda_titulo = ws["A1"]
    celda_titulo.value     = "Histórico USD/MXN — Últimos 30 días"
    celda_titulo.fill      = _fill(COLOR["title_bg"])
    celda_titulo.font      = _font(COLOR["title_fg"], bold=True, size=13)
    celda_titulo.alignment = _centro()
    ws.row_dimensions[1].height = 28

    # ── Estadísticas rápidas en fila 2
    if not df.empty:
        max_val  = df["usd_fix"].max()
        min_val  = df["usd_fix"].min()
        prom_val = df["usd_fix"].mean()

        stats = [
            ("Máximo del período", f"${max_val:,.4f}"),
            ("Mínimo del período", f"${min_val:,.4f}"),
            ("Promedio del período", f"${prom_val:,.4f}"),
        ]
        col_stat = 1
        for label, val in stats:
            _aplicar_celda(ws, 2, col_stat,     label, bg=COLOR["header_bg"], fg=COLOR["header_fg"], bold=True, size=10)
            _aplicar_celda(ws, 2, col_stat + 1, val,   bg=COLOR["fila_par"],  size=10)
            col_stat += 2
        # rellenar col E vacía
        ws.cell(row=2, column=5).fill = _fill(COLOR["fila_par"])

    # ── Encabezados tabla
    encabezados = ["Fecha", "USD/MXN (FIX)", "Var. diaria", "Máx. mes", "Mín. mes"]
    anchos      = [14, 16, 14, 14, 14]

    for col_idx, (enc, ancho) in enumerate(zip(encabezados, anchos), start=1):
        _aplicar_celda(
            ws, 3, col_idx, enc,
            bg=COLOR["header_bg"], fg=COLOR["header_fg"],
            bold=True, size=11
        )
        ws.column_dimensions[get_column_letter(col_idx)].width = ancho

    # ── Filas de datos
    for i, row in enumerate(df.itertuples(), start=4):
        bg_fila = COLOR["fila_par"] if i % 2 == 0 else COLOR["fila_impar"]

        # Fecha
        fecha_val = row.fecha.strftime("%d/%m/%Y") if pd.notna(row.fecha) else ""
        _aplicar_celda(ws, i, 1, fecha_val, bg=bg_fila)

        # USD/MXN
        _aplicar_celda(ws, i, 2, row.usd_fix, bg=bg_fila, formato='$#,##0.4f')

        # Variación diaria con color (invertido: USD sube = peso cae = malo)
        var_val = row.var_diaria if pd.notna(row.var_diaria) else None
        if var_val is not None:
            bg_var, fg_var = _color_variacion(-var_val)
        else:
            bg_var, fg_var = COLOR["neutro_bg"], COLOR["neutro_fg"]
        _aplicar_celda(
            ws, i, 3,
            _formatear_variacion(var_val),
            bg=bg_var, fg=fg_var, bold=(var_val is not None)
        )

        # Máx y mín del mes
        _aplicar_celda(ws, i, 4, row.max_mes, bg=bg_fila, formato='$#,##0.4f')
        _aplicar_celda(ws, i, 5, row.min_mes, bg=bg_fila, formato='$#,##0.4f')

        ws.row_dimensions[i].height = 18

    ws.freeze_panes = "A4"


# ── Función principal ──────────────────────────────────────────────────────

def generar_excel(df_resumen: pd.DataFrame, df_historico: pd.DataFrame) -> str:
    """
    Genera el archivo .xlsx con las dos hojas y lo guarda en outputs/.

    Retorna la ruta absoluta del archivo generado.
    """
    # Crear carpeta outputs/ si no existe
    outputs_dir = Path(__file__).parent.parent / "outputs"
    outputs_dir.mkdir(exist_ok=True)

    fecha_hoy  = date.today()
    fecha_str  = fecha_hoy.strftime("%d de %B de %Y").capitalize()
    nombre_archivo = f"reporte_economico_{fecha_hoy.strftime('%Y-%m-%d')}.xlsx"
    ruta_archivo   = outputs_dir / nombre_archivo

    print(f"Generando Excel: {nombre_archivo}...")

    wb = Workbook()

    # ── Hoja 1: Resumen
    ws_resumen = wb.active
    ws_resumen.title = "Resumen del día"
    _construir_hoja_resumen(ws_resumen, df_resumen, fecha_str)

    # ── Hoja 2: Histórico
    ws_historico = wb.create_sheet("Histórico USD")
    if not df_historico.empty:
        _construir_hoja_historico(ws_historico, df_historico)
    else:
        ws_historico["A1"] = "No hay datos históricos disponibles."

    wb.save(str(ruta_archivo))
    print(f"Excel guardado en: {ruta_archivo}")

    return str(ruta_archivo)


# ── Ejecución directa ──────────────────────────────────────────────────────

if __name__ == "__main__":
    import sys
    sys.path.append(str(Path(__file__).parent))

    from extraccion import obtener_datos
    from procesamiento import procesar_datos

    print("── Extrayendo datos...")
    datos_crudos = obtener_datos()

    print("\n── Procesando datos...")
    df_resumen, df_historico = procesar_datos(datos_crudos)

    print("\n── Generando Excel...")
    ruta = generar_excel(df_resumen, df_historico)

    print(f"\nArchivo listo: {ruta}")